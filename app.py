import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# --------------------
# 模組 1：解除合併儲存格並填入原值
# --------------------
def unmerge_and_fill(ws):
    for merged in list(ws.merged_cells.ranges):
        value = ws.cell(merged.min_row, merged.min_col).value
        ws.unmerge_cells(str(merged))
        for row in ws[merged.coord]:
            for cell in row:
                cell.value = value

# --------------------
# 模組 2：整理班表資料（含空白列）
# --------------------
def consolidate_selected_sheets(wb, sheet_names):
    all_data = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        unmerge_and_fill(ws)
        clinic_name = str(ws.cell(row=1, column=1).value)[:4]
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(1, max_row + 1):
            for c in range(2, max_col + 1):
                cell_value = ws.cell(r, c).value
                if isinstance(cell_value, datetime):
                    date_val = cell_value
                    i = r + 3
                    while i <= max_row:
                        shift_type = str(ws.cell(i, c).value).strip()
                        if isinstance(ws.cell(i, c).value, datetime) or shift_type == "":
                            break
                        if shift_type in ["早", "午", "晚"]:
                            i += 1
                            while i <= max_row:
                                if isinstance(ws.cell(i, c).value, datetime):
                                    break
                                val = str(ws.cell(i, c).value).strip()
                                if val in ["早", "午", "晚"]:
                                    break
                                all_data.append([
                                    clinic_name,
                                    date_val.strftime("%Y/%m/%d"),
                                    shift_type,
                                    val,
                                    ws.cell(i, 1).value,
                                    ws.cell(i, 21).value
                                ])
                                i += 1
                            i -= 1
                        i += 1
    df = pd.DataFrame(all_data, columns=["診所", "日期", "班別", "姓名", "A欄資料", "U欄資料"])
    return df

# --------------------
# 模組 3：建立班別分析表（含班別代碼）並過濾無效姓名
# --------------------
def create_shift_analysis(df_shift: pd.DataFrame, df_emp: pd.DataFrame) -> pd.DataFrame:
    df_shift = df_shift.copy()
    df_emp = df_emp.copy()
    df_shift.columns = [str(c).strip() for c in df_shift.columns]
    df_emp.columns = [str(c).strip() for c in df_emp.columns]

    # 自動匹配欄位
    shift_col_map = {}
    for col in df_shift.columns:
        if "姓名" in col:
            shift_col_map["name"] = col
        elif "班別" in col:
            shift_col_map["shift"] = col
        elif "診所" in col:
            shift_col_map["clinic"] = col
        elif "日期" in col:
            shift_col_map["date"] = col
        elif "A欄" in col or "E欄" in col:
            shift_col_map["e_value"] = col

    emp_col_map = {}
    for col in df_emp.columns:
        if "姓名" in col:
            emp_col_map["name"] = col
        elif "員工編號" in col or "ID" in col:
            emp_col_map["id"] = col
        elif "部門" in col:
            emp_col_map["dept"] = col
        elif "職稱" in col:
            emp_col_map["title"] = col

    # 建立員工字典
    emp_dict = {}
    for _, row in df_emp.iterrows():
        name = str(row.get(emp_col_map.get("name"), "")).strip()
        if name:
            emp_id = str(row.get(emp_col_map.get("id"), "")).strip()
            dept = str(row.get(emp_col_map.get("dept"), "")).strip()
            title = str(row.get(emp_col_map.get("title"), "")).strip()
            emp_dict[name] = (emp_id, dept, title)

    # 合併班別
    shift_dict = {}
    for _, row in df_shift.iterrows():
        name = str(row.get(shift_col_map.get("name"), "")).strip()
        clinic = str(row.get(shift_col_map.get("clinic"), "")).strip()
        date_val = row.get(shift_col_map.get("date"), "")
        shift_type = str(row.get(shift_col_map.get("shift"), "")).strip()
        e_value = row.get(shift_col_map.get("e_value"), "")

        if not name or len(name) > 4:
            continue

        key = f"{name}|{date_val}|{clinic}"
        if key not in shift_dict:
            shift_dict[key] = shift_type
        else:
            shift_dict[key] += " " + shift_type

    # 生成整理後 DataFrame
    data_out = []
    for key, shifts in shift_dict.items():
        name, date_val, clinic = key.split("|")
        shift_type = format_shift_order(shifts)
        emp_info = emp_dict.get(name, ("", "", ""))
        emp_id, emp_dept, emp_title = emp_info
        class_code = get_class_code(emp_title, clinic, shift_type)
        data_out.append([clinic, emp_id, emp_dept, name, emp_title, date_val, shift_type, e_value, class_code])

    df_analysis = pd.DataFrame(
        data_out,
        columns=["診所", "員工編號", "所屬部門", "姓名", "職稱", "日期", "班別", "E欄資料", "班別代碼"]
    )

    # 過濾無效姓名
    invalid_names = ["None", "nan", "義診", "單診", "盤點", "電打"]
    df_analysis = df_analysis[~df_analysis["姓名"].astype(str).str.strip().isin(invalid_names)].copy()

    return df_analysis

def format_shift_order(shift_str: str) -> str:
    result = ""
    for s in ["早", "午", "晚"]:
        if s in shift_str:
            result += s
    return result

def get_class_code(emp_title: str, clinic_name: str, shift_type: str) -> str:
    class_code = ""
    emp_title = str(emp_title).strip()
    if not emp_title:
        return ""

    if emp_title in ["早班護理師", "早班內視鏡助理", "醫務專員", "兼職早班內視鏡助理"]:
        return "【員工】純早班"

    if emp_title == "醫師":
        class_code = "★醫師★"
    elif emp_title in ["櫃臺", "護理師", "兼職護理師", "兼職跟診助理", "副店長", "護士", "藥師"]:
        class_code = "【員工】"
    elif "副店長" in emp_title:
        class_code = "【員工】"
    elif "店長" in emp_title or "採購儲備組長" in emp_title:
        class_code = "◇主管◇"

    if shift_type != "早":
        if clinic_name in ["上吉診所", "立吉診所", "上承診所", "立全診所", "立竹診所", "立順診所", "上京診所"]:
            class_code += "板土中京"
        elif clinic_name == "立丞診所":
            class_code += "立丞"

    shift_map = {
        "早": "早班", "午晚": "午晚班", "早午晚": "全天班",
        "早晚": "早晚班", "午": "午班", "晚": "晚班", "早午": "早午班"
    }
    class_code += shift_map.get(shift_type, shift_type)

    if class_code.endswith("早班早班"):
        class_code = class_code.replace("早班早班", "早班")
    return class_code

# --------------------
# 模組 4：建立班別總表
# --------------------
def create_shift_summary(df_analysis: pd.DataFrame) -> pd.DataFrame:
    if df_analysis.empty:
        return pd.DataFrame()

    df_analysis["日期"] = pd.to_datetime(df_analysis["日期"])
    min_date = df_analysis["日期"].min()
    max_date = df_analysis["日期"].max()
    all_dates = pd.date_range(min_date, max_date).strftime("%Y-%m-%d").tolist()

    summary_dict = {}
    for _, row in df_analysis.iterrows():
        emp_id = str(row["員工編號"])
        emp_name = row["姓名"]
        if not emp_name or str(emp_name).strip() in ["None", "nan"]:
            continue
        shift_date = row["日期"].strftime("%Y-%m-%d")
        class_code = row["班別代碼"]

        key = (emp_id, emp_name)
        if key not in summary_dict:
            summary_dict[key] = {}
        summary_dict[key][shift_date] = class_code

    data_out = []
    for (emp_id, emp_name), shifts in summary_dict.items():
        row = [emp_id, emp_name] + [shifts.get(d, "") for d in all_dates]
        data_out.append(row)

    columns = ["員工編號", "員工姓名"] + all_dates
    df_summary = pd.DataFrame(data_out, columns=columns)
    return df_summary

# --------------------
# Streamlit 主程式
# --------------------
st.title("上吉醫療-班表轉換鋒形格式")

shift_file = st.file_uploader("上傳班表 Excel 檔案", type=["xlsx", "xlsm"])
employee_file = st.file_uploader("上傳員工資料 Excel 檔案", type=["xlsx", "xlsm"])

if shift_file and employee_file:
    wb_shift = load_workbook(shift_file)
    wb_emp = load_workbook(employee_file)

    # 選擇工作表
    selectable_sheets = [s for s in wb_shift.sheetnames if s not in ["彙整結果", "班別分析", "班別總表"]]
    selected_sheets = st.multiselect("選擇要處理的工作表", selectable_sheets)
    employee_sheet_name = st.selectbox("選擇員工資料工作表", wb_emp.sheetnames)

    if st.button("開始處理"):
        if not selected_sheets:
            st.warning("請至少選擇一個工作表！")
        else:
            df_shift = consolidate_selected_sheets(wb_shift, selected_sheets)
            ws_emp = wb_emp[employee_sheet_name]
            data_emp = ws_emp.values
            cols_emp = [str(c).strip() for c in next(data_emp)]
            df_emp = pd.DataFrame(data_emp, columns=cols_emp)

            df_analysis = create_shift_analysis(df_shift, df_emp)
            df_summary = create_shift_summary(df_analysis)

            st.success("處理完成！")
            st.subheader("班別總表（已過濾無效姓名）")
            st.dataframe(df_summary)

            # 下載 Excel（只含班別總表）
            with BytesIO() as output:
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_summary.to_excel(writer, sheet_name="班別總表", index=False)
                st.download_button(
                    "下載班別總表 Excel",
                    data=output.getvalue(),
                    file_name="班別總表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
